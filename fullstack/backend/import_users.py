"""一次性账号导入：读取 Excel 模板，把学生账号 upsert 进库。

用法：
    python import_users.py                 # 默认读 templates/user_template.xlsx
    python import_users.py 路径.xlsx       # 指定文件

说明：
    本平台为封闭账号体系——账号只能由管理员通过此表格统一下发，
    不开放学生自助注册（/api/auth/register 返回 403）。
    表格列头：username | password | display_name | email(选填) | role(选填, 默认student)
    已存在的 username 会被跳过（不会覆盖原密码）。
    role 可选值：student / teacher / admin。
"""
import os
import sys
from openpyxl import load_workbook
from core.database import SessionLocal, Base, engine
import models  # noqa: F401  确保模型已注册
from core.security import hash_password

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "templates", "user_template.xlsx")


def import_users(path: str):
    Base.metadata.create_all(bind=engine)
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("表格为空。")
        return
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    col_user = idx.get("username")
    col_pwd = idx.get("password")
    col_name = idx.get("display_name") or idx.get("name")
    col_email = idx.get("email")
    col_role = idx.get("role")
    if col_user is None or col_pwd is None:
        print("缺少必要列：username / password。请使用模板列头。")
        return

    db = SessionLocal()
    created = 0
    skipped = 0
    try:
        for r in rows[1:]:
            if r is None:
                continue
            username = str(r[col_user]).strip() if r[col_user] is not None else ""
            password = str(r[col_pwd]).strip() if r[col_pwd] is not None else ""
            if not username or not password:
                continue
            name = str(r[col_name]).strip() if (col_name is not None and r[col_name]) else username
            email = str(r[col_email]).strip() if (col_email is not None and r[col_email]) else None
            role = str(r[col_role]).strip().lower() if (col_role is not None and r[col_role]) else "student"
            if role not in ("student", "teacher", "admin"):
                role = "student"
            existing = db.query(models.User).filter(models.User.username == username).first()
            if existing:
                skipped += 1
                continue
            u = models.User(
                username=username,
                name=name,
                password_hash=hash_password(password),
                avatar=name[:1],
                role=role,
                email=email,
            )
            db.add(u)
            created += 1
        db.commit()
        print(f"导入完成：新建 {created} 个账号，跳过(已存在) {skipped} 个。")
    finally:
        db.close()


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(target):
        print(f"未找到表格：{target}\n请先填写 templates/user_template.xlsx 后运行。")
        sys.exit(1)
    import_users(target)
